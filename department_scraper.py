import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time


# Initialize the Selenium WebDriver
def init_driver():
    driver_path = r'C:\Users\user\Desktop\btk\chromedriver-win64\chromedriver-win64\chromedriver.exe'  # Replace with the path to your ChromeDriver
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service)
    return driver


# Load the Excel file containing department names
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    departments = [sheet.cell(row=i, column=1).value for i in
                   range(2, sheet.max_row + 1)]  # Assuming department names start at row 2
    return departments, wb, sheet  # Returning departments, workbook, and sheet


# Search for the department in the input field
def search_department(driver, department_name):
    driver.get("https://yokatlas.yok.gov.tr/netler.php")

    # Wait for the dropdown toggle button to appear and click it
    toggle_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "btn.dropdown-toggle.btn-default"))
    )
    toggle_button.click()

    # Wait for the search box to appear after clicking the dropdown button
    search_box = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "bs-searchbox"))
    )

    # Find the input element within the 'bs-searchbox' div and enter the department name
    input_box = search_box.find_element(By.TAG_NAME, "input")
    input_box.send_keys(department_name)

    # Wait for the dropdown menu to appear and click the corresponding department
    dropdown_menu = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, "dropdown-menu.inner"))
    )

    # Find the corresponding department in the dropdown list and click it
    options = dropdown_menu.find_elements(By.TAG_NAME, "a")

    department_selected = False
    for option in options:
        if department_name == option.text:
            option.click()
            department_selected = True
            break

    if not department_selected:
        print(f"Department '{department_name}' not found in dropdown menu.")

    # Wait for the page to load after the department is selected
    time.sleep(0.2)


# Scrape the required information from the table (first row, 6th column)
def scrape_data(driver):
    try:
        # Get the page source from the Selenium driver
        page_source = driver.page_source

        # Parse the page source with BeautifulSoup
        soup = BeautifulSoup(page_source, 'html.parser')

        # Find the table with id "mydata"
        table = soup.find('table', id="mydata")

        if table:
            # Find all rows (tr) in the table
            rows = table.find_all('tr')

            # Ensure there are enough rows
            if len(rows) > 2:
                # Collect the values from the 6th column (excluding the first two rows)
                values = []
                for row in rows[2:]:
                    columns = row.find_all('td')
                    if len(columns) >= 6:
                        value = columns[5].get_text(strip=True)
                        try:
                            # Convert value to float and add to values list
                            values.append(float(value.replace(',', '.')))
                        except ValueError:
                            # Handle cases where the value is not a number
                            print(f"Non-numeric value found: {value}")
                            continue

                # Calculate the average of the collected values
                if values:
                    average = sum(values) / len(values)
                    # Return the average rounded to 3 decimal places
                    return round(average, 3)
                else:
                    print("No numeric values found for averaging.")
                    return None
            else:
                print("Not enough rows found in the table.")
                return None
        else:
            print("Table with id 'mydata' not found.")
            return None
    except Exception as e:
        print(f"Error scraping data: {e}")
        return None


# Write the department and its data to the Excel file
def write_to_excel(sheet, row, department, data):
    sheet.cell(row=row, column=1).value = department
    sheet.cell(row=row, column=2).value = data


# Main function
def main(input_excel_file, output_excel_file):
    departments, wb, sheet = load_excel(input_excel_file)  # Load the input Excel file
    driver = init_driver()

    for idx, department in enumerate(departments, start=2):  # Start at row 2
        search_department(driver, department)
        data = scrape_data(driver)

        if data:
            print(f"Data for {department}: {data}")
        else:
            print(f"No data found for {department}.")
            data = "N/A"  # Mark as not available if no data is found

        # Write the department and data to the Excel sheet
        write_to_excel(sheet, idx, department, data)

    # Save the updated Excel file with scraped data
    wb.save(output_excel_file)
    driver.quit()


if __name__ == "__main__":
    input_excel_file = "departments.xlsx"  # Path to your input Excel file
    output_excel_file = "department_data_2.xlsx"  # Path to the output Excel file
    main(input_excel_file, output_excel_file)
