import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time


# Load the Excel file
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    universities = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row+1)]  # Assuming university names start at row 2
    return universities, wb, sheet



# Initialize the Selenium driver
def init_driver():
    # Path to your webdriver executable
    driver_path = r'C:\Users\user\Desktop\btk\chromedriver-win64\chromedriver-win64\chromedriver.exe'
    service = Service(driver_path)
    driver = webdriver.Chrome(service=service)
    return driver


# Search for university name + "ortalama puan yök atlas"
def search_university(driver, university_name):
    search_query = f"{university_name} bilgisayar mühendisliği ortalama puan yök atlas"
    #search_query = f"BAHÇEŞEHİR ÜNİVERSİTESİ bilgisayar mühendisliği ortalama puan yök atlas"

    driver.get("https://www.google.com")
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(search_query)
    search_box.send_keys(Keys.RETURN)
    time.sleep(2)  # Wait for the results to load


# Check if the first result is from 'yokatlas.yok.gov.tr' and open it
def check_and_open_yokatlas(driver):
    try:
        results = driver.find_elements(By.CSS_SELECTOR, "div.yuRUbf a")  # Get all result links

        if len(results) == 0:
            print("No results found.")
            return False

        # Loop through results and find the first URL that contains 'lisans.php'
        for result in results:
            result_url = result.get_attribute("href")
            if "lisans.php" in result_url:
                print(f"Opening result with 'lisans.php' in URL: {result_url}")
                driver.get(result_url)
                return True

        print("No URL containing 'lisans.php' found in the results.")
        return False
    except Exception as e:
        print(f"Error checking YÖK Atlas link: {e}")
        return False


def close_modal_if_present(driver):
    try:
        close_button = driver.find_element(By.CLASS_NAME, "featherlight-close")
        close_button.click()
        time.sleep(0.3)  # Give time for the modal to close
    except Exception as e:
        print(f"No modal found or error closing modal: {e}")

# Click on the element that reveals the hidden table (element with id "h1220")
def click_to_reveal_table(driver):
    try:
        close_modal_if_present(driver)
        reveal_button = driver.find_element(By.ID, "h1220")
        reveal_button.click()
        time.sleep(2)  # Give time for the table to become visible
        return True
    except Exception as e:
        print(f"Error clicking the reveal button: {e}")
        return False


# Scrape the YKS score from the div with id "icerik_1220"
def scrape_yks_score(driver):
    time.sleep(2)  # Ensure the page is fully loaded
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # Locate the div with id "icerik_1220"
    div = soup.find("div", {"id": "icerik_1220"})

    if div:
        # Locate the table within that div
        table = div.find("table")
        if table:
            rows = table.find_all("tr")

            # Iterate over the rows and look for the one containing "YKS"
            for row in rows:
                cols = row.find_all("td")
                print(cols)
                if len(cols) > 0 and "TYT" in cols[0].text:
                    print(cols[1].text)
                    # Return the corresponding score for YKS
                    return cols[1].text.strip()  # Assuming the score is in the second column
    return None


# Write the university name and its score to the output Excel sheet
def write_to_excel(sheet, row, university, score):
    sheet.cell(row=row, column=1).value = university
    sheet.cell(row=row, column=2).value = score

# Main function
def main(input_excel_file, output_excel_file):
    universities, wb, sheet = load_excel(input_excel_file)  # Load the input Excel file
    driver = init_driver()

    for idx, university in enumerate(universities, start=2):  # Start at row 2
        search_university(driver, university)
        if check_and_open_yokatlas(driver):
            if click_to_reveal_table(driver):  # Click to reveal the table
                score = scrape_yks_score(driver)
                if score:
                    print(f"YKS Score for {university}: {score}")
                else:
                    print(f"No YKS score found for {university}.")
                    score = "N/A"  # Mark as not available if no score is found
            else:
                print(f"Unable to reveal table for {university}.")
                score = "N/A"
        else:
            print(f"No valid YÖK Atlas page found for {university}.")
            score = "N/A"

        # Write the university and score to the Excel sheet
        write_to_excel(sheet, idx, university, score)

    # Save the updated Excel file with scores
    wb.save(output_excel_file)
    driver.quit()

if __name__ == "__main__":
    input_excel_file = "university_list_test.xlsx"  # Path to your input Excel file
    output_excel_file = "university_scores_2.xlsx"  # Path to the output Excel file
    main(input_excel_file, output_excel_file)